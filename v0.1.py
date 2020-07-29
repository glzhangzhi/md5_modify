#! python3
from os import chdir, walk
from os.path import abspath, dirname, exists, relpath

from hashlib import md5
from openpyxl import Workbook, load_workbook


def check():
    with open('result.txt', 'w') as f:
        wb = load_workbook('data.xlsx')
        ws = wb['Sheet1']
        for cell_c in ws['C']:
            if cell_c.value is None:
                break
            tag = True
            for cell_a in ws['A']:
                if cell_a.value == cell_c.value:
                    tag = False
                    if ws['B' + str(cell_a.row)].value != ws['D' + str(cell_c.row)].value:
                        f.write(cell_c.value + '   修改' + '\n')
            if tag:
                f.write(cell_c.value + '   新增' + '\n')
            else:
                pass

        for cell_a in ws['A']:
            tag = True
            if cell_a.value is None:
                break
            for cell_c in ws['C']:
                if cell_a.value == cell_c.value:
                    tag = False
            if tag:
                f.write(cell_a.value + '   删除' + '\n')


def collect(now_path):
    if exists('data.xlsx'):
        wb = load_workbook('data.xlsx')
        letter = 'C'
        tag = True
    else:
        wb = Workbook()
        wb.create_sheet(index=0, title='Sheet1')
        letter = 'A'
        tag = False
    ws = wb['Sheet1']
    n = 1
    for foldername, subfolders, filenames in walk(now_path):
        for filename in filenames:
            file_name = foldername + '\\' + filename
            if filename == 'test.py' or filename == 'data.xlsx':
                continue
            else:
                pass
            ws[letter + str(n)] = relpath(file_name)
            with open(file_name, 'rb') as fp:
                data = fp.read()
            file_md5 = md5(data).hexdigest()
            ws[chr(ord(letter) + 1) + str(n)] = file_md5
            n += 1
    wb.save('data.xlsx')
    if tag:
        check()


def main():
    now_path = dirname(abspath(__file__))
    chdir(now_path)
    collect(now_path)


if __name__ == '__main__':
    main()
